import os

import requests
from bs4 import BeautifulSoup
import json

from openpyxl import load_workbook


def getjson(year=2015,month=3):

    url = "https://tianqi.2345.com/Pc/GetHistory"
    params = {
        # "areaInfo[areaId]": "58362",
        "areaInfo[areaId]": "71072",
        "areaInfo[areaType]": "2",
        "date[year]": year,
        "date[month]": month
    }
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "X-Requested-With": "XMLHttpRequest",
        "Referer": "https://tianqi.2345.com/wea_history/58362.htm"
    }

    response = requests.get(url, params=params, headers=headers)

    if response.status_code == 200:
        # print("成功获取数据：")
        # 根据实际情况解析json或其他格式的数据
        # print(response.json())
        return response.json()
    else:
        print(f"请求失败，状态码：{response.status_code}")


def toAnanysi(response):
    html_content = response['data']

    # 解析HTML内容
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取平均高温和平均低温
    avg_high_temp = soup.find('em', class_='orange-txt').text if soup.find('em', class_='orange-txt') else 'N/A'
    avg_low_temp = soup.find_all('em', class_='blue-txt')[1].text if len(
        soup.find_all('em', class_='blue-txt')) > 1 else 'N/A'

    # 提取极端高温及其日期
    extreme_high_temp = soup.find_all('em', class_='orange-txt')[1].text if len(
        soup.find_all('em', class_='orange-txt')) > 1 else 'N/A'
    extreme_high_date = soup.find_all('b')[0].text.strip('()') if soup.find_all('b') else 'N/A'

    # 提取极端低温及其日期
    extreme_low_temp = soup.find_all('em', class_='blue-txt')[0].text if len(
        soup.find_all('em', class_='blue-txt')) > 0 else 'N/A'
    extreme_low_date = soup.find_all('b')[1].text.strip('()') if len(soup.find_all('b')) > 1 else 'N/A'

    # 提取表格中的数据
    table_data = []
    table = soup.find('table', class_='history-table')
    if table:
        rows = table.find_all('tr')
        for row in rows[1:]:  # 跳过表头
            cols = row.find_all('td')
            date = cols[0].text.strip()
            high_temp = cols[1].text.strip()
            low_temp = cols[2].text.strip()
            weather = cols[3].text.strip()
            wind = cols[4].text.strip()
            table_data.append({
                '日期': date,
                '最高温': high_temp,
                '最低温': low_temp,
                '天气': weather,
                '风力风向': wind
            })

    # 构建输出数据结构
    weather_data = {
        "average_temperature": {"high": avg_high_temp, "low": avg_low_temp},
        "extreme_temperature": {
            "high": {"temperature": extreme_high_temp, "date": extreme_high_date},
            "low": {"temperature": extreme_low_temp, "date": extreme_low_date}
        },
        "daily_records": table_data
    }

    # 将结果转换为JSON格式并打印
    json_output = json.dumps(weather_data, ensure_ascii=False, indent=4)
    print(json_output)

import json
from bs4 import BeautifulSoup
import pandas as pd

def toanalysis(response):
    html_content = response['data']

    # 解析HTML内容
    soup = BeautifulSoup(html_content, 'html.parser')

    # 提取平均高温和平均低温
    avg_high_temp = soup.find('em', class_='orange-txt').text if soup.find('em', class_='orange-txt') else 'N/A'
    avg_low_temp = soup.find_all('em', class_='blue-txt')[1].text if len(
        soup.find_all('em', class_='blue-txt')) > 1 else 'N/A'

    # 提取极端高温及其日期
    extreme_high_temp = soup.find_all('em', class_='orange-txt')[1].text if len(
        soup.find_all('em', class_='orange-txt')) > 1 else 'N/A'
    extreme_high_date = soup.find_all('b')[0].text.strip('()') if soup.find_all('b') else 'N/A'

    # 提取极端低温及其日期
    extreme_low_temp = soup.find_all('em', class_='blue-txt')[0].text if len(
        soup.find_all('em', class_='blue-txt')) > 0 else 'N/A'
    extreme_low_date = soup.find_all('b')[1].text.strip('()') if len(soup.find_all('b')) > 1 else 'N/A'

    # 提取表格中的数据
    table_data = []
    table = soup.find('table', class_='history-table')
    if table:
        rows = table.find_all('tr')
        for row in rows[1:]:  # 跳过表头
            cols = row.find_all('td')
            date = cols[0].text.strip()
            high_temp = cols[1].text.strip()
            low_temp = cols[2].text.strip()
            weather = cols[3].text.strip()
            wind = cols[4].text.strip()
            table_data.append({
                'date': date,
                'high_temp': high_temp,
                'low_temp': low_temp,
                'weather': weather,
                'wind': wind
            })

    # 构建输出数据结构
    weather_data = {
        "average_temperature": {"high": avg_high_temp, "low": avg_low_temp},
        "extreme_temperature": {
            "high": {"temperature": extreme_high_temp, "date": extreme_high_date},
            "low": {"temperature": extreme_low_temp, "date": extreme_low_date}
        },
        "daily_records": table_data
    }

    # 将结果转换为JSON格式并打印7
    json_output = json.dumps(weather_data, ensure_ascii=False, indent=4)
    # 将 JSON 字符串解析回 Python 字典
    datas = json.loads(json_output)
    return datas

def toExcel(df, file_path="data.xlsx", sheet_name="Sheet1", reset=False,):
    """
    @rest重置表格sheet数据参数
    print(f"数据已写入 {file_path} 的 {sheet_name}，{'重置' if reset else '追加'}模式")
    """
    if reset or not os.path.exists(file_path):
        # 覆盖写入（新建文件或重置）
        with pd.ExcelWriter(file_path, mode='w', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        if os.path.exists(file_path):
        # 加载已有数据
            with pd.ExcelWriter(file_path, mode='a', engine='openpyxl', if_sheet_exists='overlay') as writer:
                book = load_workbook(file_path)
                writer._book = book
                if sheet_name in writer.book.sheetnames:
                    start_row = writer.book[sheet_name].max_row
                else:
                    start_row = 0  # 如果 sheet 不存在，则从头写入
                    # df_daily_records.to_excel(writer, sheet_name=sheet_name, index=False)
                df.to_excel(writer, sheet_name=sheet_name, index=False, header=(start_row == 0), startrow=start_row)

                # df.to_excel(writer, sheet_name=sheet_name, index=False, header=False,
                #                         startrow=writer.sheets[sheet_name].max_row)
        else:
       # 创建新文件
            with pd.ExcelWriter(file_path, mode='w', engine='openpyxl') as writer:
                df_daily_records.to_excel(writer, sheet_name=sheet_name, index=False)
if __name__ == "__main__":
    jsondata=getjson(year=2025,month=3)
    pandasdata = toanalysis(jsondata)

    # 将 daily_records 转换为 DataFrame
    try:
        df_daily_records = pd.DataFrame(pandasdata['daily_records'])
        toExcel(df_daily_records, file_path="data.xlsx", sheet_name="sheet1", reset=False)
        print(df_daily_records)
    except TypeError as e:
        print(f"遇到类型错误: {e}")
        print("请检查输入的数据是否与预期相符。")

